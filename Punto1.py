from openpyxl import*
from gurobipy import*
from prettytable import*
import networkx as nx
import matplotlib.pyplot as plt
# Carga de Parametros
book = load_workbook("Tarea 2-202420.xlsx")
Sheet=book["Opti-drones"]


#----------------------------------------------
# Definicion de Conjuntos
#----------------------------------------------

# Conjunto de Lugares
L = [] 

for fila in range(2,28):
    lugar = Sheet.cell(fila,1).value
    L.append(lugar)

# Conjunto de Arcos posibles
A = [] 
for fila1 in range(2,28):
    lugar1 = Sheet.cell(fila1,1).value
    
    for fila2 in range(2,28):
        lugar2 = Sheet.cell(fila2,1).value
        if lugar1 != lugar2:
            lugar = (lugar1, lugar2)
            A.append(lugar)



#----------------------------------------------
# Definicion de Parametros
#----------------------------------------------

# Posicion X (Carrera)
X = {}
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posX =  Sheet.cell(fila,2).value
    X[lugar] = posX

# Posicion Y (Calle)
Y = {} 
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posY =  Sheet.cell(fila,3).value
    Y[lugar] = posY


# Cantidad de fotos requeridas por lugar
F = {}
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posY =  Sheet.cell(fila,4).value
    F[lugar] = posY


# Distancia Manhattan entre lugares
M = {}
for i in L:
    for j in L:
        if i != j: #Para que no se repitan los lugares
            M[(i,j)] = abs(X[i]-X[j])+abs(Y[i]-Y[j])

#----------------------------------------------
# Modelo de Optimización
#----------------------------------------------

#Creación del modelo
m = Model("Tarea2")

#Definicion de variables de decisión
x = m.addVars(A,vtype=GRB.BINARY, name = "x")

#----------------------------------------------
# Restricciones del problema 
#----------------------------------------------


#1
for i in L:
    if i != 'Hangar':
        m.addConstr(quicksum(x[i,j] for j in L if i != j) == 1)

for j in L:
    if j != 'Hangar':
        m.addConstr(quicksum(x[i,j] for i in L if i != j) == 1)
        
#2
m.addConstr(quicksum(x['Hangar',j] for j in L if j != 'Hangar') == 5)
m.addConstr(quicksum(x[i,'Hangar'] for i in L if i != 'Hangar') == 5)


#3
for i in L :
    for j in L :
        if i !=j: 
            m.addConstr(x[i,j] + x[j,i] <= 1)

#F.O
m.setObjective(quicksum(x[a]*M[a] for a in A),GRB.MINIMIZE)
m.update()
#m.setParam("Outputflag",0)
m.optimize()
arcos_reales = []
for i in A:
    if x[i].x > 0:
        arcos_reales.append(i)


def tiempo_secuencia(secuencia):
        tiempo = 0
        for i in range(len(secuencia)-1):
            tiempo += (M[(secuencia[i], secuencia[i+1])]/10)
            if secuencia[i+1] != 'Hangar':
                tiempo += 90
        return round(tiempo / 60, 2)  # redondear a dos decimales
    
    
    
def cantidad_fotos(secuencia):
        fotos = 0
        for i in range(len(secuencia)-1):
            fotos += F[secuencia[i]]
        return fotos


#Funcion para graficar y mostrar tabla
def graficar_tabla(arcos_reales):
    G = nx.DiGraph()
    G.add_edges_from(arcos_reales)
    nx.draw(G, with_labels=True)
    ciclos = list(nx.simple_cycles(G))
    cantidad_ciclos = len(ciclos)
    table = PrettyTable()
    table.field_names = ["#Ruta", "Secuencia", "Tiempo (Horas)", "Fotos"]
    
    #Agregar el primer nodo al final de cada ciclo
    for i in range(cantidad_ciclos):
        ciclos[i].append(ciclos[i][0])
    
    for i in range(cantidad_ciclos):
        table.add_row([i+1,ciclos[i],tiempo_secuencia(ciclos[i]) , cantidad_fotos(ciclos[i])])
    print(table)
    plt.show()
    
graficar_tabla(arcos_reales)

#Para los ciclos que superan tiempo de 12 horas, crear restricciones
#----------------------------------------------
# Restricciones del problema
#----------------------------------------------
G = nx.DiGraph()
G.add_edges_from(arcos_reales)
ciclos = list(nx.simple_cycles(G))
cantidad_ciclos = len(ciclos)


#Agregar el primer nodo al final de cada ciclo
for i in range(cantidad_ciclos):
    ciclos[i].append(ciclos[i][0])
    
#Funcion para hacer una lista de los nodos de una secuencia
def nodos_ciclo(ciclo):
    nodos = []
    for i in range(len(ciclo)-1):
        nodos.append(ciclo[i])
    return nodos

#Funcion para hacer una lista de los arcos de un ciclo
def arcos_ciclo(ciclo):
    arcos = []
    for i in range(len(ciclo)-1):
        arcos.append((ciclo[i],ciclo[i+1]))
    return arcos
        
print(nodos_ciclo(ciclos[2]))
print(arcos_ciclo(ciclos[2]))

print(quicksum(M[a]/600 + 1.5 for a in arcos_ciclo(ciclos[2])))



#for i in range(len(ciclos)):
 #   if tiempo_secuencia(ciclos[i]) > 12:
# Calculate the total time of the cycle that exceeds the limit
ciclo_2 = ciclos[2]
tiempo_ciclo_2 = tiempo_secuencia(ciclo_2)

# If the cycle exceeds 12 hours, add a constraint to the model
#if tiempo_ciclo_2 > 12:
 #   m.addConstr(quicksum(x[ciclo_2[i], ciclo_2[i+1]] * M[(ciclo_2[i], ciclo_2[i+1])] for i in range(len(ciclo_2)-1)) <= 720, "Restriccion_Tiempo_Ciclo_2")


#m.addConstr(quicksum(x[a]*M[a]/600 + 1.5 for a in arcos_ciclo(ciclos[2])) <= 12)

#Funcion para eliminar subrutas que no pasan por el Hangar
def eliminar_subrutas_sin_hangar(modelo, ciclos, M, x):
    while True:
        subrutas_sin_hangar = []
        # Recalcula los ciclos actuales
        for idx, ciclo in enumerate(ciclos):
            if 'Hangar' not in ciclo:  # Verifica si el ciclo no contiene el "Hangar"
                print(f"Ciclo {idx+1} no pasa por el Hangar: {ciclo}")
                subrutas_sin_hangar.append(ciclo)

        if not subrutas_sin_hangar:
            print("Todas las rutas pasan por el Hangar.")
            break

        # Agregar restricciones para eliminar las subrutas que no pasan por el Hangar
        for ciclo in subrutas_sin_hangar:
            arcos = arcos_ciclo(ciclo)  # Obtener los arcos del ciclo
            # Restricción para eliminar la subruta
            modelo.addConstr(quicksum(x[a] for a in arcos) <= len(arcos) - 1)
            print(f"Agregando restricción para eliminar subruta: {ciclo}")

        # Reoptimizamos después de agregar las nuevas restricciones
        modelo.update()
        modelo.optimize()

        # Recalcular los ciclos reales tras la optimización
        arcos_reales = []
        for i in A:
            if x[i].x > 0:
                arcos_reales.append(i)
        
        # Recalcula los ciclos reales usando la librería NetworkX
        G = nx.DiGraph()
        G.add_edges_from(arcos_reales)
        ciclos = list(nx.simple_cycles(G))  # Obtener los ciclos de la nueva solución
        for i in range(len(ciclos)):
            ciclos[i].append(ciclos[i][0])  # Cerrar los ciclos (volver al hangar)








#Funcion para restringir ciclos que superan 12 horas
def verificar_y_restringir_ciclos(modelo, ciclos, M, x):
    while True:
        ciclos_a_restringir = []
        # Recalcula los ciclos actuales
        for idx, ciclo in enumerate(ciclos):
            tiempo_ciclo = tiempo_secuencia(ciclo)  # Usa tu función para calcular el tiempo del ciclo
            if tiempo_ciclo > 12:
                print(f"Ciclo {idx+1} excede las 12 horas: {tiempo_ciclo} horas")
                ciclos_a_restringir.append(ciclo)

        if not ciclos_a_restringir:
            print("Todos los ciclos cumplen con el límite de tiempo de 12 horas.")
            break

        # Agregar restricciones para los ciclos que exceden las 12 horas
        for ciclo in ciclos_a_restringir:
            arcos = arcos_ciclo(ciclo)  # Obtener los arcos del ciclo
            modelo.addConstr(quicksum(x[a] * (M[a] / 600 + (1.5 if a[1] != 'Hangar' else 0)) for a in arcos) <= 12)
            print(f"Agregando restricción para ciclo: {ciclo}")

        # Reoptimizamos después de agregar las nuevas restricciones
        modelo.update()
        modelo.optimize()

        # Recalcular los ciclos para ver si hay alguno que aún excede las 12 horas
        arcos_reales = []
        for i in A:
            if x[i].x > 0:
                arcos_reales.append(i)
        
        # Recalcula los ciclos reales
        G = nx.DiGraph()
        G.add_edges_from(arcos_reales)
        ciclos = list(nx.simple_cycles(G))  # Obtener los ciclos de la nueva solución
        for i in range(len(ciclos)):
            ciclos[i].append(ciclos[i][0])  # Cerrar los ciclos (volver al hangar)


# Ejecutar el algoritmo iterativo



#Funcion para restringir ciclos que superan las 300 fotos
def verificar_y_restringir_fotos(modelo, ciclos, F, x):
    while True:
        ciclos_a_restringir = []
        # Recalcula los ciclos actuales
        for idx, ciclo in enumerate(ciclos):
            fotos_ciclo = cantidad_fotos(ciclo)  # Usa tu función para calcular la cantidad de fotos del ciclo
            if fotos_ciclo > 300:
                print(f"Ciclo {idx+1} excede las 300 fotos: {fotos_ciclo} fotos")
                ciclos_a_restringir.append(ciclo)

        if not ciclos_a_restringir:
            print("Todos los ciclos cumplen con el límite de 300 fotos.")
            break

        # Agregar restricciones para los ciclos que exceden las 300 fotos
        for ciclo in ciclos_a_restringir:
            modelo.addConstr(quicksum(x[i,j]*F[i] for i in nodos_ciclo(ciclo) for j in nodos_ciclo(ciclo)) <= 300)
            print(f"Agregando restricción para ciclo: {ciclo}")

        # Reoptimizamos después de agregar las nuevas restricciones
        modelo.update()
        modelo.optimize()

        # Recalcular los ciclos para ver si hay alguno que aún excede las 300 fotos
        arcos_reales = []
        for i in A:
            if x[i].x > 0:
                arcos_reales.append(i)
        
        # Recalcula los ciclos reales
        G = nx.DiGraph()
        G.add_edges_from(arcos_reales)
        ciclos = list(nx.simple_cycles(G))  # Obtener los ciclos de la nueva solución
        for i in range(len(ciclos)):
            ciclos[i].append(ciclos[i][0])  # Cerrar los ciclos (volver al hangar)



eliminar_subrutas_sin_hangar(m, ciclos, M, x)
verificar_y_restringir_ciclos(m, ciclos, M, x)
verificar_y_restringir_fotos(m, ciclos, F, x)










arcos_reales = []

for i in A:
    if x[i].x > 0:
        arcos_reales.append(i)
graficar_tabla(arcos_reales)



