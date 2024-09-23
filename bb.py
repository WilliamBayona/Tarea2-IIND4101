from openpyxl import load_workbook
from gurobipy import Model, GRB, quicksum
from prettytable import PrettyTable
import networkx as nx
import matplotlib.pyplot as plt

# Carga de Parametros
book = load_workbook("Tarea 2-202420.xlsx")
Sheet=book["Opti-drones"]


#----------------------------------------------
# Definicion de Conjuntos
#----------------------------------------------

# Conjunto de Lugares
L = [] 

for fila in range(2,28):
    lugar = Sheet.cell(fila,1).value
    L.append(lugar)

# Conjunto de Arcos posibles
A = [] 
for fila1 in range(2,28):
    lugar1 = Sheet.cell(fila1,1).value
    
    for fila2 in range(2,28):
        lugar2 = Sheet.cell(fila2,1).value
        if lugar1 != lugar2:
            lugar = (lugar1, lugar2)
            A.append(lugar)



#----------------------------------------------
# Definicion de Parametros
#----------------------------------------------

# Posicion X (Carrera)
X = {}
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posX =  Sheet.cell(fila,2).value
    X[lugar] = posX

# Posicion Y (Calle)
Y = {} 
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posY =  Sheet.cell(fila,3).value
    Y[lugar] = posY


# Cantidad de fotos requeridas por lugar
F = {}
for fila in range(2,28):
    lugar =  Sheet.cell(fila,1).value
    posY =  Sheet.cell(fila,4).value
    F[lugar] = posY


# Distancia Manhattan entre lugares
M = {}
for i in L:
    for j in L:
        if i != j: #Para que no se repitan los lugares
            M[(i,j)] = abs(X[i]-X[j])+abs(Y[i]-Y[j])

#----------------------------------------------
# Modelo de Optimización
#----------------------------------------------

#Creación del modelo
m = Model("Tarea2")

#Definicion de variables de decisión
x = m.addVars(A,vtype=GRB.BINARY, name = "x")

#----------------------------------------------
# Restricciones del problema 
#----------------------------------------------


#1. Todos los lugares deben ser visitados por un dron
for i in L:
    if i != 'Hangar':
        m.addConstr(quicksum(x[i,j] for j in L if i != j) == 1)

for j in L:
    if j != 'Hangar':
        m.addConstr(quicksum(x[i,j] for i in L if i != j) == 1)
        
#2. Restricción para que los drones salgan y lleguen al hangar
m.addConstr(quicksum(x['Hangar',j] for j in L if j != 'Hangar') == 5)
m.addConstr(quicksum(x[i,'Hangar'] for i in L if i != 'Hangar') == 5)


#3. Restricción para evitar ciclos de dos nodos
for i in L :
    for j in L :
        if i !=j: 
            m.addConstr(x[i,j] + x[j,i] <= 1)

#F.O
m.setObjective(quicksum(x[a]*M[a] for a in A),GRB.MINIMIZE)
m.update()
m.setParam("Outputflag",0)
m.optimize()

#---------------------------
# resultados parte A
#--------------------------




#----------------------------------------------
# Funciones auxiliares
#----------------------------------------------

def tiempo_secuencia(secuencia):
    tiempo = 0
    for i in range(len(secuencia)-1):
        tiempo += (M[(secuencia[i], secuencia[i+1])]/10)
        if secuencia[i+1] != 'Hangar':
            tiempo += 90
    return round(tiempo / 60, 2)  

def cantidad_fotos(secuencia):
    fotos = 0
    for i in range(len(secuencia)-1):
        fotos += F[secuencia[i]]
    return fotos

def graficar_tabla(arcos_reales):
    G = nx.DiGraph()
    G.add_edges_from(arcos_reales)
    nx.draw(G, with_labels=True)
    ciclos = list(nx.simple_cycles(G))
    cantidad_ciclos = len(ciclos)
    table = PrettyTable()
    table.field_names = ["#Ruta", "Secuencia", "Tiempo (Horas)", "Fotos"]
    
    for i in range(cantidad_ciclos):
        ciclos[i].append(ciclos[i][0])
    
    for i in range(cantidad_ciclos):
        table.add_row([i+1,ciclos[i],tiempo_secuencia(ciclos[i]) , cantidad_fotos(ciclos[i])])
    print(table)
    plt.savefig("final.png")
    plt.show()
    
def nodos_ciclo(ciclo):
    nodos = []
    for i in range(len(ciclo)-1):
        nodos.append(ciclo[i])
    return nodos
#---------------------------
# resultados parte A
#--------------------------

arcos_reales = [(i, j) for i, j in A if x[i, j].x > 0]
G = nx.DiGraph()
G.add_edges_from(arcos_reales)
ciclos = list(nx.simple_cycles(G))
cantidad_ciclos = len(ciclos)
table = PrettyTable()
table.field_names = ["#Ruta", "Secuencia", "Tiempo (Horas)", "Fotos"]

for i in range(cantidad_ciclos):
    ciclos[i].append(ciclos[i][0])

for i in range(cantidad_ciclos):
    table.add_row([i+1,ciclos[i],tiempo_secuencia(ciclos[i]) , cantidad_fotos(ciclos[i])])
FO = m.getObjective().getValue()
print(FO)
print(table)


#----------------------------------------------
# Funciones Punto B.
#----------------------------------------------

valoresFO = {}

def verificar_y_restringir(modelo, arcos_reales, M,F, x):
    """Restringe ciclos que exceden 12 horas,300 fotos y que no pasan por el hangar"""
    contador = 0
    while True:
        G = nx.DiGraph()
        G.add_edges_from(arcos_reales)
        ciclos = list(nx.simple_cycles(G))
    
        nx.draw(G, with_labels = True)
        plt.savefig("filename" + str(contador) +".png")
        plt.clf()
        for i in range(len(ciclos)):
            ciclos[i].append(ciclos[i][0])
            


#Restriccion de 300 fotos
#----------------------------------------------
        ciclos_a_restringir2 = [ciclo for ciclo in ciclos if cantidad_fotos(ciclo) > 300]
        for ciclo in ciclos_a_restringir2:
            modelo.addConstr(quicksum(x[i, j] * F[i] for i in nodos_ciclo(ciclo) for j in nodos_ciclo(ciclo) if i != j) <= 300)
            print(f"Agregando restricción para ciclo: {ciclo}")
#----------------------------------------------  

#Restriccion de 12 horas
#----------------------------------------------
        ciclos_a_restringir1 = [ciclo for ciclo in ciclos if tiempo_secuencia(ciclo) > 12]      
        for ciclo in ciclos_a_restringir1:
            arcos = [(ciclo[i], ciclo[i + 1]) for i in range(len(ciclo) - 1)]
            modelo.addConstr(quicksum(x[a] * (M[a] / 600 + (1.5 if a[1] != 'Hangar' else 0)) for a in arcos) <= 12)
            print(f"Agregando restricción para ciclo: {ciclo}")
#----------------------------------------------

#Restriccion de Hangar
#----------------------------------------------
        subrutas_sin_hangar = [ciclo for ciclo in ciclos if 'Hangar' not in ciclo]
        for ciclo in subrutas_sin_hangar:           
            modelo.addConstr(quicksum(x[i, j] for i in ciclo for j in L if i != 'Hangar' and j not in ciclo)>= 1)
            print(f"Agregando restricción para eliminar subruta: {ciclo}")
#----------------------------------------------
        
        if not subrutas_sin_hangar and not ciclos_a_restringir1 and not ciclos_a_restringir2:
            break     

        modelo.update()
        modelo.optimize()
        contador += 1
        FO = m.getObjective().getValue()
        valoresFO[contador] = FO
        
        arcos_reales = [(i, j) for i, j in A if x[i, j].x > 0] 



#----------------------------------------------
# Aplicar restricciones iterativamente
#----------------------------------------------

arcos_reales = [(i, j) for i, j in A if x[i, j].x > 0]

while True:
    # Guardar la solución actual
    
    solucion_anterior = [(i, j) for i, j in A if x[i, j].x > 0]
    
    # Aplicar las funciones de restricción
    verificar_y_restringir(m, arcos_reales,M, F, x)

    # Verificar si la solución ha cambiado
    arcos_reales = [(i, j) for i, j in A if x[i, j].x > 0] 
    if arcos_reales == solucion_anterior:
        break

# Mostrar la solución final
print("Funcion Objetivo" , m.getObjective().getValue())
graficar_tabla(arcos_reales)
print(valoresFO)
#graficar la tabla
plt.plot(valoresFO.keys(), valoresFO.values())
plt.xlabel("Iteraciones")
plt.ylabel("Valor de la Función Objetivo")
plt.title("Valor de la Función Objetivo por Iteración")
plt.savefig("Comparativa.png")
plt.show()

#----------------------------------------------